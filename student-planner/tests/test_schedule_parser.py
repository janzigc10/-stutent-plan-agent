from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

import openpyxl

from app.services.schedule_parser import RawCourse, parse_excel_schedule

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "sample_schedule.xlsx"
GAOSHU = "\u9ad8\u7b49\u6570\u5b66"
XIANDAI = "\u7ebf\u6027\u4ee3\u6570"
YINGYU = "\u5927\u5b66\u82f1\u8bed"
GAILV = "\u6982\u7387\u8bba"
TIYU = "\u4f53\u80b2"
ZHANG = "\u5f20\u8001\u5e08"
LI = "\u674e\u8001\u5e08"
WANG = "\u738b\u8001\u5e08"
ZHAO = "\u8d75\u8001\u5e08"
TEACHING_A301 = "\u6559\u5b66\u697cA301"
TEACHING_B205 = "\u6559\u5b66\u697cB205"
FOREIGN_201 = "\u5916\u8bed\u697c201"
TEACHING_A302 = "\u6559\u5b66\u697cA302"
PLAYGROUND = "\u64cd\u573a"


def test_parse_returns_list_of_raw_courses() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    assert isinstance(courses, list)
    assert len(courses) > 0
    assert all(isinstance(course, RawCourse) for course in courses)


def test_parse_extracts_course_name() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    names = {course.name for course in courses}
    assert GAOSHU in names
    assert XIANDAI in names
    assert YINGYU in names


def test_parse_extracts_weekday() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == GAOSHU)
    assert gaoshu.weekday == 1


def test_parse_extracts_period() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == GAOSHU)
    assert gaoshu.period == "1-2"


def test_parse_extracts_teacher() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == GAOSHU)
    assert gaoshu.teacher == ZHANG


def test_parse_extracts_location() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == GAOSHU)
    assert gaoshu.location == TEACHING_A301


def test_parse_extracts_weeks() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == GAOSHU)
    assert gaoshu.week_start == 1
    assert gaoshu.week_end == 16


def test_parse_handles_missing_teacher() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    tiyu = next(course for course in courses if course.name == TIYU)
    assert tiyu.teacher is None


def test_parse_handles_custom_week_range() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gailv = next(course for course in courses if course.name == GAILV)
    assert gailv.week_start == 3
    assert gailv.week_end == 16


def test_parse_total_course_count() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    assert len(courses) == 6


def test_parse_falls_back_to_xls_parser_when_openpyxl_rejects_stream(
    monkeypatch,
) -> None:
    expected = [
        RawCourse(
            name=GAOSHU,
            teacher=ZHANG,
            location=TEACHING_A301,
            weekday=1,
            period="1-2",
            week_start=1,
            week_end=16,
        )
    ]

    def fake_parse_xlsx_workbook(_file_path):
        raise BadZipFile("File is not a zip file")

    def fake_parse_xls_bytes(file_bytes: bytes):
        assert file_bytes == b"legacy-xls-bytes"
        return expected

    monkeypatch.setattr("app.services.schedule_parser._parse_xlsx_workbook", fake_parse_xlsx_workbook)
    monkeypatch.setattr("app.services.schedule_parser._parse_xls_bytes", fake_parse_xls_bytes)

    courses = parse_excel_schedule(BytesIO(b"legacy-xls-bytes"))
    assert courses == expected


def test_parse_detects_header_not_in_first_row() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "schedule"
    sheet["A1"] = "2025-2026 schedule"
    sheet["A3"] = "period"
    sheet["B3"] = "monday"
    sheet["C3"] = "tuesday"
    sheet["A4"] = "1-2"
    sheet["B4"] = f"{GAOSHU}\n{ZHANG}\n{TEACHING_A301}\n1-16\u5468"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)
    assert len(courses) == 1
    assert courses[0].name == GAOSHU
    assert courses[0].weekday == 1
    assert courses[0].period == "1-2"


def test_parse_uses_sheet_with_most_courses() -> None:
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "notes"
    first["A1"] = "this is not a schedule"

    second = workbook.create_sheet(title="schedule")
    second["A1"] = "period"
    second["B1"] = "wednesday"
    second["A2"] = "3-4"
    second["B2"] = f"{YINGYU}\n{LI}\n{FOREIGN_201}\n1-16\u5468"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)
    assert len(courses) == 1
    assert courses[0].name == YINGYU
    assert courses[0].weekday == 3


def test_parse_cell_with_blank_line_keeps_single_course_block() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "course-table"
    sheet["A1"] = "period"
    sheet["B1"] = "friday"
    sheet["A2"] = "7-8"
    sheet["B2"] = f"{TIYU}\n\n{PLAYGROUND}\n1-16\u5468"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)
    assert len(courses) == 1
    assert courses[0].name == TIYU
    assert courses[0].location == PLAYGROUND
    assert courses[0].period == "7-8"


def test_parse_supports_odd_even_and_single_week_expressions() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "course-table"
    sheet["A1"] = "period"
    sheet["B1"] = "monday"
    sheet["C1"] = "tuesday"
    sheet["D1"] = "wednesday"
    sheet["A2"] = "1-2"
    sheet["B2"] = "odd-course\nTeacher A\nA301\n\u7b2c1-18\u5468(\u5355\u5468)"
    sheet["C2"] = "even-course\nTeacher B\nB201\n\u7b2c2-16\u5468(\u53cc\u5468)"
    sheet["D2"] = "single-week-course\nTeacher C\nC101\n\u7b2c1\u5468"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)

    odd_week = next(course for course in courses if course.name == "odd-course")
    even_week = next(course for course in courses if course.name == "even-course")
    single_week = next(course for course in courses if course.name == "single-week-course")

    assert odd_week.week_start == 1
    assert odd_week.week_end == 18
    assert odd_week.week_pattern == "odd"
    assert odd_week.week_text == "\u7b2c1-18\u5468(\u5355\u5468)"

    assert even_week.week_start == 2
    assert even_week.week_end == 16
    assert even_week.week_pattern == "even"
    assert even_week.week_text == "\u7b2c2-16\u5468(\u53cc\u5468)"

    assert single_week.week_start == 1
    assert single_week.week_end == 1
    assert single_week.week_pattern == "all"
    assert single_week.week_text == "\u7b2c1\u5468"
